#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""创建 Excel 映射模板文件"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def create_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "文件发送映射"

    # 样式
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="07C160", end_color="07C160", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # 表头
    headers = ["文件名关键词", "微信备注名/昵称", "备注（可选）"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 示例数据
    examples = [
        ("月报", "张三", "文件名包含【月报】即匹配"),
        ("工资单", "李四", "如: 2024年08月工资单.xlsx"),
        ("合同", "王五", "如: 劳动合同-张三.pdf"),
        ("发票", "赵六", "如: 增值税发票_001.pdf"),
    ]
    for row, (keyword, name, note) in enumerate(examples, 2):
        for col, val in enumerate([keyword, name, note], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # 说明行
    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value="使用说明:").font = Font(name="Microsoft YaHei", size=10, bold=True)
    notes = [
        "1. 第一列填文件名关键词，第二列填对方的微信备注名（优先）或昵称",
        "2. 匹配规则: 文件名包含关键词即匹配（不区分大小写）",
        "3. 一个关键词匹配到多个文件时，全部发送给对应联系人",
        "4. 多个关键词映射到同一个人时会自动合并去重，只发送一次",
        "5. 第一行为表头，程序自动跳过",
        "6. 请确保微信PC版已登录；重名时建议填微信号（唯一）",
    ]
    for i, note in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=note).font = Font(name="Microsoft YaHei", size=9, color="888888")

    # 列宽
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40

    # 保存
    output = r"C:\Users\10129\WorkBuddy\2026-08-03-10-51-00\发送映射模板.xlsx"
    wb.save(output)
    print(f"模板已保存: {output}")

if __name__ == "__main__":
    create_template()
