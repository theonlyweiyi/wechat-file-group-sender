#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
微信文件群发工具 — Web 版
==========================
Flask + 现代化 Web UI，浏览器打开即用。

启动后访问: http://localhost:5890
"""

import os
import sys
import time
import json
import threading
from datetime import datetime
from io import BytesIO

from flask import Flask, request, jsonify, send_from_directory, send_file, Response
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from wechat_sender import WeChatSender, set_log_fn as sender_set_log

app = Flask(__name__)
app.secret_key = 'wechat_sender_secret_key_2026'

UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '_uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

TEMPLATE_NAME = "发送映射模板.xlsx"

def _resolve_template_source():
    """定位内嵌的模板资源；打包后从 sys._MEIPASS 读取，否则从脚本目录读取。"""
    candidates = []
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        candidates.append(os.path.join(sys._MEIPASS, TEMPLATE_NAME))
    candidates.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), TEMPLATE_NAME))
    for path in candidates:
        if os.path.isfile(path):
            return path
    return None

def _build_template_bytes():
    """动态生成模板 xlsx 的字节流（兜底方案，保证单文件零附带）。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "文件发送映射"

    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="07C160", end_color="07C160", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Microsoft YaHei", size=10)
    cell_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    headers = ["文件名关键词", "微信昵称", "备注（可选）"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    examples = [
        ("月报", "张三", "文件名包含【月报】即匹配"),
        ("工资单", "李四", "如: 2024年08月工资单.xlsx"),
        ("合同", "王五", "如: 劳动合同-张三.pdf"),
        ("发票", "赵六", "如: 增值税发票_001.pdf"),
    ]
    for row, (keyword, name, note) in enumerate(examples, 2):
        for col, val in enumerate([keyword, name, note], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    note_row = len(examples) + 3
    ws.cell(row=note_row, column=1, value="使用说明:").font = Font(name="Microsoft YaHei", size=10, bold=True)
    notes = [
        "1. 第一列填文件名关键词，第二列填微信昵称（与微信显示一致）",
        "2. 匹配规则: 文件名包含关键词即匹配（不区分大小写）",
        "3. 一个关键词匹配到多个文件时，全部发送给对应联系人",
        "4. 第一行为表头，程序自动跳过",
        "5. 请确保微信PC版已登录，昵称/备注名与微信一致",
    ]
    for i, note in enumerate(notes, 1):
        ws.cell(row=note_row + i, column=1, value=note).font = Font(name="Microsoft YaHei", size=9, color="888888")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── 全局状态 ───────────────────────────────────────────────
state = {
    "folder_path": "",
    "excel_uploaded": None,        # (filename, path)
    "mappings": [],                 # [(keyword, name), ...]
    "matched": [],                  # [(keyword, name, [files]), ...]
    "sending": False,
    "stop_requested": False,
    "logs": [],
    "progress": {"current": 0, "total": 0},
    "result": None,                 # "completed" / "stopped" / None
}

# ── 日志 ──────────────────────────────────────────────────

def _add_log(msg):
    """向日志列表添加条目"""
    ts = datetime.now().strftime("%H:%M:%S")
    state["logs"].append({"time": ts, "msg": msg})
    # 保留最近 200 条
    if len(state["logs"]) > 200:
        state["logs"] = state["logs"][-200:]


# ══════════════════════════════════════════════════════════════
#  前端页面
# ══════════════════════════════════════════════════════════════

HTML = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>微信文件群发工具</title>
<style>
  :root {
    --bg: #f0f2f5;
    --card: #fff;
    --primary: #07c160;
    --primary-hover: #06ad56;
    --danger: #e74c3c;
    --text: #1f2937;
    --text-light: #6b7280;
    --text-muted: #9ca3af;
    --border: #e5e7eb;
    --input-bg: #f9fafb;
    --shadow: 0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.06);
    --radius: 10px;
    --radius-sm: 6px;
  }

  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
  }

  .container { max-width: 860px; margin: 0 auto; padding: 24px 20px; }

  /* Header */
  .header { text-align: center; padding: 32px 0; }
  .header h1 { font-size: 26px; font-weight: 700; color: var(--text); }
  .header p { color: var(--text-light); margin-top: 6px; font-size: 14px; }

  /* Card */
  .card {
    background: var(--card);
    border-radius: var(--radius);
    box-shadow: var(--shadow);
    padding: 24px;
    margin-bottom: 16px;
  }
  .card-title {
    font-size: 15px; font-weight: 600; color: var(--text);
    margin-bottom: 16px; display: flex; align-items: center; gap: 8px;
  }
  .card-title .icon { font-size: 18px; }

  /* Form */
  .form-row { display: flex; gap: 10px; align-items: flex-end; flex-wrap: wrap; }
  .form-group { flex: 1; min-width: 200px; }
  .form-group label { display: block; font-size: 13px; font-weight: 600; color: var(--text-light); margin-bottom: 6px; }
  .form-group input, .form-group textarea {
    width: 100%; padding: 10px 14px; font-size: 14px;
    border: 1.5px solid var(--border); border-radius: var(--radius-sm);
    background: var(--input-bg); color: var(--text);
    transition: border-color .15s; outline: none; font-family: inherit;
  }
  .form-group input:focus, .form-group textarea:focus { border-color: var(--primary); box-shadow: 0 0 0 3px rgba(7,193,96,.1); }
  .form-group textarea { resize: vertical; min-height: 56px; }
  .form-group .hint { font-size: 12px; color: var(--text-muted); margin-top: 4px; }

  /* Buttons */
  .btn {
    display: inline-flex; align-items: center; gap: 6px;
    padding: 10px 20px; font-size: 14px; font-weight: 600;
    border: none; border-radius: var(--radius-sm); cursor: pointer;
    transition: all .15s; font-family: inherit;
    white-space: nowrap;
  }
  .btn:disabled { opacity: .5; cursor: not-allowed; }
  .btn-primary { background: var(--primary); color: #fff; }
  .btn-primary:hover:not(:disabled) { background: var(--primary-hover); }
  .btn-outline { background: #fff; color: var(--text); border: 1.5px solid var(--border); }
  .btn-outline:hover:not(:disabled) { background: #f9fafb; border-color: var(--primary); color: var(--primary); }
  .btn-danger { background: var(--danger); color: #fff; }
  .btn-danger:hover:not(:disabled) { background: #c0392b; }
  .btn-sm { padding: 6px 14px; font-size: 13px; }
  .btn-block { width: 100%; justify-content: center; }

  /* Badge */
  .badge {
    display: inline-block; padding: 2px 10px; font-size: 12px; font-weight: 600;
    border-radius: 20px; line-height: 1.6;
  }
  .badge-success { background: #d1fae5; color: #065f46; }
  .badge-warn { background: #fef3c7; color: #92400e; }
  .badge-fail { background: #fee2e2; color: #991b1b; }
  .badge-info { background: #dbeafe; color: #1e40af; }

  /* File list */
  .file-list {
    max-height: 180px; overflow-y: auto; margin-top: 8px;
    font-size: 13px; color: var(--text-light);
    padding: 8px 12px; background: var(--input-bg);
    border-radius: var(--radius-sm); border: 1px solid var(--border);
  }
  .file-list .file-item { padding: 3px 0; display: flex; align-items: center; gap: 6px; }
  .file-list .file-icon { color: var(--primary); }

  /* Table */
  .table-wrap { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { text-align: left; padding: 10px 12px; font-weight: 600; color: var(--text-light); border-bottom: 2px solid var(--border); font-size: 12px; text-transform: uppercase; letter-spacing: .5px; }
  td { padding: 10px 12px; border-bottom: 1px solid var(--border); vertical-align: top; }
  tr:hover td { background: #f9fafb; }
  .file-list-inline { font-size: 12px; color: var(--text-light); }
  .file-list-inline div { padding: 1px 0; }
  .no-data { text-align: center; color: var(--text-muted); padding: 40px 0; font-size: 14px; }

  /* Tabs */
  .tabs { display: flex; gap: 0; margin-bottom: 16px; border-bottom: 2px solid var(--border); }
  .tab {
    padding: 8px 20px; font-size: 13px; font-weight: 600; cursor: pointer;
    border: none; background: none; color: var(--text-muted);
    border-bottom: 2px solid transparent; margin-bottom: -2px;
    transition: all .15s; font-family: inherit;
  }
  .tab.active { color: var(--primary); border-bottom-color: var(--primary); }
  .tab:hover { color: var(--text); }

  /* Manual input table */
  .mapping-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .mapping-table th { padding: 8px 10px; text-align: left; font-size: 12px; }
  .mapping-table td { padding: 4px 6px; border-bottom: 1px solid var(--border); }
  .mapping-table input {
    width: 100%; padding: 8px 10px; font-size: 13px;
    border: 1.5px solid var(--border); border-radius: var(--radius-sm);
    background: var(--input-bg); color: var(--text); outline: none;
    font-family: inherit;
  }
  .mapping-table input:focus { border-color: var(--primary); }
  .btn-icon {
    background: none; border: none; cursor: pointer; color: var(--text-muted);
    font-size: 16px; padding: 4px 8px; border-radius: var(--radius-sm);
    transition: all .15s;
  }
  .btn-icon:hover { color: var(--danger); background: #fee2e2; }
  .add-row-btn { margin-top: 10px; }

  /* Logs */
  .log-area {
    background: #1e1e1e; color: #d4d4d4; border-radius: var(--radius-sm);
    padding: 12px 16px; font-family: "Cascadia Code", "Fira Code", Consolas, monospace;
    font-size: 12px; line-height: 1.7; max-height: 260px; overflow-y: auto;
  }
  .log-line { white-space: pre-wrap; word-break: break-all; }
  .log-line.success { color: #4ecb71; }
  .log-line.fail { color: #f14c4c; }
  .log-line.info { color: #569cd6; }
  .log-line.warn { color: #cca700; }

  /* Progress */
  .progress-wrap { margin: 14px 0; }
  .progress-bar {
    width: 100%; height: 8px; background: var(--border);
    border-radius: 4px; overflow: hidden;
  }
  .progress-fill {
    height: 100%; background: var(--primary);
    border-radius: 4px; transition: width .3s;
    width: 0%;
  }
  .progress-label {
    display: flex; justify-content: space-between;
    font-size: 12px; color: var(--text-light); margin-top: 6px;
  }

  /* Upload zone */
  .upload-zone {
    border: 2px dashed var(--border); border-radius: var(--radius-sm);
    padding: 28px; text-align: center; cursor: pointer;
    transition: all .15s; color: var(--text-muted);
  }
  .upload-zone:hover, .upload-zone.drag-over {
    border-color: var(--primary); color: var(--primary); background: rgba(7,193,96,.03);
  }
  .upload-zone .upload-icon { font-size: 32px; margin-bottom: 8px; }
  .upload-zone .upload-text { font-size: 14px; }
  .upload-zone .upload-hint { font-size: 12px; margin-top: 4px; }
  .upload-zone input[type="file"] { display: none; }

  .template-link { margin-top: 10px; text-align: center; }
  .btn-text {
    background: none; border: none; cursor: pointer;
    color: var(--brand, #07c160); font-size: 13px; padding: 4px 6px;
    text-decoration: underline; font-family: inherit;
  }
  .btn-text:hover { color: #05984c; }

  .file-uploaded {
    display: flex; align-items: center; gap: 8px;
    padding: 8px 14px; background: #d1fae5; border-radius: var(--radius-sm);
    font-size: 13px; color: #065f46; margin-top: 8px;
  }

  /* Toggle */
  .toggle-row { display: flex; align-items: center; gap: 10px; margin-bottom: 10px; }
  .toggle-row label { font-size: 13px; font-weight: 600; color: var(--text-light); cursor: pointer; }

  /* Toast */
  .toast {
    position: fixed; top: 20px; left: 50%; transform: translateX(-50%);
    padding: 12px 24px; border-radius: var(--radius-sm);
    font-size: 14px; font-weight: 600; color: #fff; z-index: 9999;
    animation: toastIn .3s ease;
  }
  .toast.success { background: var(--primary); }
  .toast.error { background: var(--danger); }
  @keyframes toastIn { from { opacity: 0; transform: translateX(-50%) translateY(-10px); } to { opacity: 1; transform: translateX(-50%) translateY(0); } }

  /* Responsive */
  @media (max-width: 640px) {
    .container { padding: 12px; }
    .card { padding: 16px; }
    .form-row { flex-direction: column; }
  }
</style>
</head>
<body>
<div class="container">

  <div class="header">
    <h1>📁 微信文件群发工具</h1>
    <p>按关键词匹配文件，自动发送给对应的微信联系人</p>
  </div>

  <!-- 步骤 1: 文件夹 -->
  <div class="card">
    <div class="card-title"><span class="icon">📂</span> 步骤 1：选择文件所在的文件夹</div>
    <div class="form-row">
      <div class="form-group">
        <label>文件夹路径</label>
        <input type="text" id="folderPath" placeholder="例如: D:\工作\8月文件">
      </div>
      <div style="padding-bottom: 2px; display: flex; gap: 8px;">
        <button class="btn btn-outline" onclick="selectFolder()">📂 浏览选择</button>
        <button class="btn btn-outline" onclick="scanFolder()" id="btnScan">🔍 检查文件夹</button>
      </div>
    </div>
    <div id="folderFiles" class="file-list" style="display:none;"></div>
  </div>

  <!-- 步骤 2: 映射表 -->
  <div class="card">
    <div class="card-title"><span class="icon">📋</span> 步骤 2：配置映射关系</div>
    <div class="tabs">
      <button class="tab active" onclick="switchMappingTab('upload')" id="tabUpload">📄 上传 Excel</button>
      <button class="tab" onclick="switchMappingTab('manual')" id="tabManual">✏️ 手动输入</button>
    </div>
    <!-- 上传模式 -->
    <div id="sectionUpload">
      <div class="upload-zone" id="uploadZone" onclick="document.getElementById('excelFile').click()">
        <div class="upload-icon">📄</div>
        <div class="upload-text">点击选择或拖拽 Excel 文件到此处</div>
        <div class="upload-hint">支持 .xlsx / .xls / .csv 格式 | 第一行为表头，自动跳过</div>
        <input type="file" id="excelFile" accept=".xlsx,.xls,.csv" onchange="handleExcelUpload(event)">
      </div>
      <div class="template-link">
        <button class="btn-text" onclick="downloadTemplate()">⬇ 没有模板？点此下载映射表模板</button>
      </div>
      <div id="excelStatus" style="display:none;"></div>
    </div>
    <!-- 手动输入模式 -->
    <div id="sectionManual" style="display:none;">
      <div class="table-wrap">
        <table class="mapping-table" id="mappingTable">
          <thead><tr><th style="width:45%">文件名关键词</th><th style="width:45%">微信昵称</th><th style="width:50px;"></th></tr></thead>
          <tbody id="mappingBody">
            <tr>
              <td><input type="text" placeholder="例如: 月报" class="kw-input"></td>
              <td><input type="text" placeholder="例如: 张三" class="name-input"></td>
              <td><button class="btn-icon" onclick="removeMappingRow(this)" title="删除">✕</button></td>
            </tr>
          </tbody>
        </table>
      </div>
      <button class="btn btn-outline btn-sm add-row-btn" onclick="addMappingRow()">➕ 添加一行</button>
      <button class="btn btn-primary btn-sm add-row-btn" onclick="saveManualMappings()" style="margin-left:8px;">💾 保存映射</button>
      <div id="manualStatus" style="display:none;margin-top:8px;"></div>
    </div>
  </div>

  <!-- 步骤 3: 消息 -->
  <div class="card">
    <div class="card-title"><span class="icon">💬</span> 步骤 3：附带消息（可选）</div>
    <div class="toggle-row">
      <input type="checkbox" id="enableMsg" onclick="toggleMessage()">
      <label for="enableMsg">发送文件时附带一条消息</label>
    </div>
    <div class="form-group" id="msgGroup" style="display:none;">
      <textarea id="msgText" placeholder="输入消息内容，支持变量: {name}=微信昵称"></textarea>
      <div class="hint">提示: 使用 <code>{name}</code> 会自动替换为对方的微信昵称</div>
    </div>
  </div>

  <!-- 操作按钮 -->
  <div style="display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap;">
    <button class="btn btn-primary" onclick="previewMatches()" id="btnPreview">🔍 预览匹配结果</button>
    <button class="btn btn-primary" onclick="startSend()" id="btnSend" disabled>🚀 开始发送</button>
    <button class="btn btn-danger" onclick="stopSend()" id="btnStop" disabled>⏹ 停止</button>
    <span id="sendStatus" style="font-size:13px;color:var(--text-light);display:flex;align-items:center;margin-left:auto;"></span>
  </div>

  <!-- 匹配预览 -->
  <div class="card" id="previewCard" style="display:none;">
    <div class="card-title"><span class="icon">📊</span> 匹配预览</div>
    <div class="table-wrap">
      <table>
        <thead><tr><th>关键词</th><th>微信昵称</th><th>匹配文件</th><th style="width:70px;">状态</th></tr></thead>
        <tbody id="previewBody"></tbody>
      </table>
    </div>
    <div class="no-data" id="previewEmpty">暂无匹配数据，请先点击「预览匹配结果」</div>
  </div>

  <!-- 进度 -->
  <div id="progressArea" style="display:none;">
    <div class="progress-wrap">
      <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
      <div class="progress-label">
        <span id="progressText">0 / 0</span>
        <span id="progressPercent">0%</span>
      </div>
    </div>
  </div>

  <!-- 日志 -->
  <div class="card">
    <div class="card-title"><span class="icon">📝</span> 发送日志</div>
    <div class="log-area" id="logArea">
      <div class="log-line info">就绪，等待操作...</div>
    </div>
  </div>

</div>

<script>
// ─── 映射表模式切换 ──────────────────────────────────────────
let mappingMode = 'upload';  // 'upload' | 'manual'

async function switchMappingTab(mode) {
  mappingMode = mode;
  document.getElementById('sectionUpload').style.display = mode === 'upload' ? 'block' : 'none';
  document.getElementById('sectionManual').style.display = mode === 'manual' ? 'block' : 'none';
  document.getElementById('tabUpload').classList.toggle('active', mode === 'upload');
  document.getElementById('tabManual').classList.toggle('active', mode === 'manual');
}

function addMappingRow() {
  const tr = document.createElement('tr');
  tr.innerHTML = `
    <td><input type="text" placeholder="例如: 月报" class="kw-input"></td>
    <td><input type="text" placeholder="例如: 张三" class="name-input"></td>
    <td><button class="btn-icon" onclick="removeMappingRow(this)" title="删除">✕</button></td>`;
  document.getElementById('mappingBody').appendChild(tr);
}

function removeMappingRow(btn) {
  const tbody = document.getElementById('mappingBody');
  if (tbody.children.length <= 1) return; // 至少保留一行
  btn.closest('tr').remove();
}

async function saveManualMappings() {
  const rows = document.querySelectorAll('#mappingBody tr');
  const mappings = [];
  for (const row of rows) {
    const kw = row.querySelector('.kw-input').value.trim();
    const name = row.querySelector('.name-input').value.trim();
    if (kw && name) mappings.push([kw, name]);
  }
  if (mappings.length === 0) return toast('请至少填写一行关键词和昵称', 'error');

  const res = await fetch('/api/set_mappings', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({mappings})
  });
  const data = await res.json();
  const el = document.getElementById('manualStatus');
  el.style.display = 'block';
  el.innerHTML = `<div class="file-uploaded">📝 已保存 — ${data.count} 条映射</div>`;
  document.getElementById('excelStatus').style.display = 'none';
  toast(`已保存 ${data.count} 条映射`, 'success');
}
async function selectFolder() {
  // 优先使用 pywebview 的原生文件夹选择对话框（桌面模式）
  if (window.pywebview && window.pywebview.api && window.pywebview.api.select_folder) {
    try {
      const path = await window.pywebview.api.select_folder();
      if (path) {
        document.getElementById('folderPath').value = path;
        scanFolder();
      }
      return;
    } catch (e) { /* 走兜底 */ }
  }
  // 兜底：后端弹对话框
  const res = await fetch('/api/select_folder', { method: 'POST' });
  const data = await res.json();
  if (data.path) {
    document.getElementById('folderPath').value = data.path;
    scanFolder();
  }
}

// ─── 文件夹扫描 ────────────────────────────────────────────
async function scanFolder() {
  const path = document.getElementById('folderPath').value.trim();
  if (!path) return toast('请先输入文件夹路径', 'error');

  const res = await fetch('/api/scan_folder', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({path})
  });
  const data = await res.json();
  const el = document.getElementById('folderFiles');

  if (data.error) {
    el.style.display = 'block';
    el.innerHTML = `<span style="color:var(--danger);">⚠ ${escapeHtml(data.error)}</span>`;
    return;
  }

  el.style.display = 'block';
  if (data.files.length === 0) {
    el.innerHTML = '<span style="color:var(--text-muted);">📭 文件夹为空</span>';
  } else {
    el.innerHTML = `<div style="margin-bottom:4px;font-weight:600;">找到 ${data.count} 个文件:</div>` +
      data.files.map(f =>
        `<div class="file-item"><span class="file-icon">📄</span> ${escapeHtml(f)}</div>`
      ).join('');
  }
}

// ─── Excel 上传 ────────────────────────────────────────────
async function handleExcelUpload(event) {
  const file = event.target.files[0];
  if (!file) return;

  const formData = new FormData();
  formData.append('file', file);

  const res = await fetch('/api/upload_excel', { method: 'POST', body: formData });
  const data = await res.json();
  const el = document.getElementById('excelStatus');

  if (data.error) {
    el.style.display = 'block';
    el.innerHTML = `<div style="color:var(--danger);margin-top:8px;">⚠ ${escapeHtml(data.error)}</div>`;
  } else {
    el.style.display = 'block';
    el.innerHTML = `<div class="file-uploaded">📑 ${escapeHtml(data.filename)} — ${data.count} 条映射</div>`;
  }
}

// ─── 下载模板 ──────────────────────────────────────────────
function downloadTemplate() {
  const a = document.createElement('a');
  a.href = '/api/download_template';
  a.download = '发送映射模板.xlsx';
  document.body.appendChild(a);
  a.click();
  document.body.removeChild(a);
}

// ─── 消息开关 ──────────────────────────────────────────────
function toggleMessage() {
  document.getElementById('msgGroup').style.display =
    document.getElementById('enableMsg').checked ? 'block' : 'none';
}

// ─── 预览匹配 ──────────────────────────────────────────────
async function previewMatches() {
  const path = document.getElementById('folderPath').value.trim();
  if (!path) return toast('请先输入文件夹路径', 'error');

  const res = await fetch('/api/preview', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({path})
  });
  const data = await res.json();

  if (data.error) return toast(data.error, 'error');

  document.getElementById('previewCard').style.display = 'block';
  const tbody = document.getElementById('previewBody');
  const empty = document.getElementById('previewEmpty');

  if (data.matched.length === 0) {
    tbody.innerHTML = '';
    empty.style.display = 'block';
  } else {
    empty.style.display = 'none';
    tbody.innerHTML = data.matched.map((m, i) => {
      const filesHtml = m.files.length > 0
        ? '<div class="file-list-inline">' + m.files.map(f => `<div>📄 ${escapeHtml(f)}</div>`).join('') + '</div>'
        : '<span style="color:var(--text-muted);">—</span>';
      const badge = m.files.length > 0
        ? '<span class="badge badge-info">待发送</span>'
        : '<span class="badge badge-warn">无匹配</span>';
      return `<tr><td>${escapeHtml(m.keyword)}</td><td>${escapeHtml(m.name)}</td><td>${filesHtml}</td><td>${badge}</td></tr>`;
    }).join('');
  }

  document.getElementById('btnSend').disabled = data.sendable === 0;
  document.getElementById('sendStatus').textContent =
    `匹配: ${data.matched_total} 条 | 可发送: ${data.sendable} 个联系人 | 文件: ${data.total_files} 个`;
}

// ─── 发送 ──────────────────────────────────────────────────
async function startSend() {
  const path = document.getElementById('folderPath').value.trim();
  const enableMsg = document.getElementById('enableMsg').checked;
  const msg = document.getElementById('msgText').value;

  const res = await fetch('/api/send', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({path, enable_msg: enableMsg, msg})
  });
  const data = await res.json();
  if (data.error) return toast(data.error, 'error');

  document.getElementById('btnSend').disabled = true;
  document.getElementById('btnStop').disabled = false;
  document.getElementById('btnPreview').disabled = true;
  document.getElementById('progressArea').style.display = 'block';

  toast('已开始发送', 'success');
  pollStatus();
}

async function stopSend() {
  await fetch('/api/stop', { method: 'POST' });
  document.getElementById('btnStop').disabled = true;
  toast('正在停止...', 'error');
}

// ─── 状态轮询 ──────────────────────────────────────────────
let pollTimer = null;
async function pollStatus() {
  const res = await fetch('/api/status');
  const data = await res.json();

  // 日志
  const logEl = document.getElementById('logArea');
  logEl.innerHTML = data.logs.slice(-50).map(l => {
    let cls = '';
    if (l.msg.startsWith('✅')) cls = 'success';
    else if (l.msg.startsWith('❌')) cls = 'fail';
    else if (l.msg.startsWith('⚠')) cls = 'warn';
    else if (l.msg.startsWith('ℹ') || l.msg.includes('连接') || l.msg.includes('就绪')) cls = 'info';
    return `<div class="log-line ${cls}">[${l.time}] ${escapeHtml(l.msg)}</div>`;
  }).join('');
  logEl.scrollTop = logEl.scrollHeight;

  // 进度
  const p = data.progress;
  const pct = p.total > 0 ? Math.round(p.current / p.total * 100) : 0;
  document.getElementById('progressFill').style.width = pct + '%';
  document.getElementById('progressText').textContent = `${p.current} / ${p.total}`;
  document.getElementById('progressPercent').textContent = pct + '%';

  // 发送状态
  document.getElementById('sendStatus').textContent =
    data.sending ? `正在发送: ${p.current}/${p.total}` : '';

  if (data.sending) {
    pollTimer = setTimeout(pollStatus, 800);
  } else {
    document.getElementById('btnSend').disabled = false;
    document.getElementById('btnStop').disabled = true;
    document.getElementById('btnPreview').disabled = false;
    if (data.result === 'completed') toast('全部发送完成！', 'success');
    else if (data.result === 'stopped') toast('发送已停止', 'error');
    document.getElementById('sendStatus').textContent = data.result === 'completed' ? '✅ 全部完成' : '';
    // 发送结束，把桌面窗口焦点拉回本程序
    if (window.pywebview && window.pywebview.api && window.pywebview.api.bring_to_front) {
      try { window.pywebview.api.bring_to_front(); } catch (e) {}
    }
  }
}

// ─── 工具函数 ──────────────────────────────────────────────
function toast(msg, type) {
  const el = document.createElement('div');
  el.className = 'toast ' + type;
  el.textContent = msg;
  document.body.appendChild(el);
  setTimeout(() => el.remove(), 2500);
}
function escapeHtml(s) {
  const d = document.createElement('div');
  d.textContent = s;
  return d.innerHTML;
}

// ─── 拖拽上传 ──────────────────────────────────────────────
const uploadZone = document.getElementById('uploadZone');
uploadZone.addEventListener('dragover', e => { e.preventDefault(); uploadZone.classList.add('drag-over'); });
uploadZone.addEventListener('dragleave', () => uploadZone.classList.remove('drag-over'));
uploadZone.addEventListener('drop', e => {
  e.preventDefault();
  uploadZone.classList.remove('drag-over');
  const file = e.dataTransfer.files[0];
  if (file) {
    const dt = new DataTransfer();
    dt.items.add(file);
    document.getElementById('excelFile').files = dt.files;
    handleExcelUpload({target: {files: dt.files}});
  }
});
</script>
</body>
</html>'''


# ══════════════════════════════════════════════════════════════
#  API 路由
# ══════════════════════════════════════════════════════════════

@app.route('/')
def index():
    return HTML


@app.route('/api/select_folder', methods=['POST'])
def select_folder():
    """调用 pywebview 原生文件夹选择对话框（桌面模式下由前端直接走 JS API，这里为兜底）"""
    try:
        import webview
        if webview.windows:
            result = webview.windows[0].create_file_dialog(webview.FOLDER_DIALOG)
            if result:
                folder = os.path.normpath(result[0])
                _add_log(f"ℹ 已选择文件夹: {folder}")
                return jsonify({"path": folder})
    except Exception as e:
        _add_log(f"⚠ 文件夹选择失败: {e}")
    return jsonify({"path": ""})


@app.route('/api/set_mappings', methods=['POST'])
def set_mappings():
    """手动设置映射表（不走 Excel 上传）"""
    data = request.get_json()
    mappings_raw = data.get('mappings', [])
    mappings = [(item[0].strip(), item[1].strip()) for item in mappings_raw if len(item) == 2 and item[0].strip() and item[1].strip()]
    if not mappings:
        return jsonify({"error": "映射数据为空"})
    state["mappings"] = mappings
    state["excel_uploaded"] = ("手动输入", None)
    _add_log(f"ℹ 手动输入: {len(mappings)} 条映射")
    return jsonify({"count": len(mappings)})


@app.route('/api/scan_folder', methods=['POST'])
def scan_folder():
    data = request.get_json()
    folder = data.get('path', '').strip()
    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "文件夹不存在，请检查路径"})

    files = []
    try:
        for name in os.listdir(folder):
            full = os.path.join(folder, name)
            if os.path.isfile(full):
                files.append(name)
    except PermissionError:
        return jsonify({"error": "没有权限访问该文件夹"})

    state["folder_path"] = folder
    _add_log(f"ℹ 已选择文件夹: {folder} ({len(files)} 个文件)")
    return jsonify({"path": folder, "count": len(files), "files": sorted(files)})


@app.route('/api/upload_excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify({"error": "请选择文件"})
    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "未选择文件"})

    # secure_filename 会丢掉中文，用时间戳+扩展名保存
    ext = os.path.splitext(file.filename)[1]
    save_name = datetime.now().strftime("%Y%m%d_%H%M%S") + ext
    save_path = os.path.join(UPLOAD_FOLDER, save_name)
    file.save(save_path)

    # 解析 Excel
    try:
        mappings = _read_excel(save_path)
    except Exception as e:
        try: os.remove(save_path)
        except: pass
        return jsonify({"error": f"解析 Excel 失败: {e}"})

    state["excel_uploaded"] = (file.filename, save_path)
    state["mappings"] = mappings
    _add_log(f"ℹ 已上传映射表: {file.filename} ({len(mappings)} 条映射)")
    return jsonify({"filename": file.filename, "count": len(mappings)})


@app.route('/api/download_template')
def download_template():
    """提供映射表模板下载；优先读取编译进 exe 的资源，读取失败则用 openpyxl 动态生成。"""
    src = _resolve_template_source()
    if src:
        return send_file(src, as_attachment=True, download_name=TEMPLATE_NAME,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    buf = _build_template_bytes()
    return send_file(buf, as_attachment=True, download_name=TEMPLATE_NAME,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/api/preview', methods=['POST'])
def preview():
    data = request.get_json()
    folder = data.get('path', state.get("folder_path", "")).strip()

    if not folder or not os.path.isdir(folder):
        return jsonify({"error": "请先选择有效的文件夹路径"})
    if not state["mappings"]:
        return jsonify({"error": "请先上传 Excel 映射表"})

    # 扫描文件
    files_in_folder = {}
    for name in os.listdir(folder):
        full = os.path.join(folder, name)
        if os.path.isfile(full):
            files_in_folder[name] = full

    # 匹配
    matched = []
    sendable = 0
    total_files = 0
    for keyword, name in state["mappings"]:
        kw_lower = keyword.lower()
        file_matches = [fn for fn, fp in files_in_folder.items() if kw_lower in fn.lower()]
        matched.append({
            "keyword": keyword,
            "name": name,
            "files": file_matches,
        })
        if file_matches:
            sendable += 1
            total_files += len(file_matches)

    state["folder_path"] = folder
    state["matched"] = matched
    _add_log(f"ℹ 预览完成: {len(matched)} 条映射, {sendable} 个联系人可发送, 共 {total_files} 个文件")
    return jsonify({
        "matched": matched,
        "matched_total": len(matched),
        "sendable": sendable,
        "total_files": total_files,
    })


@app.route('/api/send', methods=['POST'])
def send():
    if state["sending"]:
        return jsonify({"error": "正在发送中，请等待完成或先停止"})
    if not state["matched"]:
        return jsonify({"error": "请先预览匹配结果"})

    data = request.get_json()
    enable_msg = data.get("enable_msg", False)
    msg = data.get("msg", "").strip()

    # 过滤出有文件的条目
    send_list = []
    for m in state["matched"]:
        if m["files"]:
            send_list.append(m)

    if not send_list:
        return jsonify({"error": "没有可发送的文件"})

    state["sending"] = True
    state["stop_requested"] = False
    state["logs"] = []
    state["progress"] = {"current": 0, "total": len(send_list)}
    state["result"] = None

    thread = threading.Thread(
        target=_send_worker,
        args=(send_list, enable_msg, msg),
        daemon=True
    )
    thread.start()

    return jsonify({"ok": True, "total": len(send_list)})


@app.route('/api/stop', methods=['POST'])
def stop():
    state["stop_requested"] = True
    _add_log("⏹ 用户请求停止...")
    return jsonify({"ok": True})


@app.route('/api/status')
def status():
    return jsonify({
        "sending": state["sending"],
        "logs": state["logs"][-100:],
        "progress": state["progress"],
        "result": state["result"],
    })


# ══════════════════════════════════════════════════════════════
#  核心逻辑
# ══════════════════════════════════════════════════════════════

def _read_excel(filepath):
    """读取 Excel/CSV 映射表"""
    ext = os.path.splitext(filepath)[1].lower()
    mappings = []

    if ext in ('.xlsx', '.xls'):
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
            if i == 0:
                continue
            if not row or len(row) < 2:
                continue
            kw = str(row[0]).strip() if row[0] else ""
            name = str(row[1]).strip() if row[1] else ""
            if kw and name:
                mappings.append((kw, name))
        wb.close()
    elif ext == '.csv':
        import csv
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i == 0:
                    continue
                if len(row) < 2:
                    continue
                kw = row[0].strip()
                name = row[1].strip()
                if kw and name:
                    mappings.append((kw, name))
    return mappings


def _send_worker(send_list, enable_msg, msg_template):
    """后台发送线程 —— 键盘模拟方案"""
    import pythoncom
    pythoncom.CoInitialize()

    total = len(send_list)

    # ── 获取文件夹内文件的全路径 ──
    folder = state["folder_path"]
    file_paths_cache = {}
    for name in os.listdir(folder):
        fp = os.path.join(folder, name)
        if os.path.isfile(fp):
            file_paths_cache[name] = fp

    # ── 连接微信 ──
    _add_log("正在连接微信客户端（键盘模拟模式）...")
    sender = WeChatSender()
    sender_set_log(_add_log)

    if not sender.connect():
        _add_log("⚠ 请确保微信PC版已登录且窗口未被最小化")
        state["sending"] = False
        state["result"] = "stopped"
        pythoncom.CoUninitialize()
        return

    _add_log("✅ 微信连接成功，开始发送...")

    success = 0
    fail = 0

    for idx, m in enumerate(send_list):
        if state["stop_requested"]:
            _add_log(f"⏹ 已停止发送 (已完成 {idx}/{total})")
            state["result"] = "stopped"
            break

        keyword, name, file_names = m["keyword"], m["name"], m["files"]
        state["progress"]["current"] = idx + 1

        try:
            # 构造消息
            message = None
            if enable_msg and msg_template:
                message = msg_template.replace("{name}", name).strip()

            # 获取文件全路径
            if len(file_names) == 1:
                file_paths = file_paths_cache[file_names[0]]
            else:
                file_paths = [file_paths_cache[fn] for fn in file_names]

            # 发送
            display = ", ".join(file_names)

            if message:
                _add_log(f"  [{idx+1}/{total}] 正在发送: {name} ← {display} [+消息]")
            else:
                _add_log(f"  [{idx+1}/{total}] 正在发送: {name} ← {display}")

            ok = sender.send(name, message=message, file_paths=file_paths)

            if ok:
                _add_log(f"✅ [{idx+1}/{total}] {name} ← {display}")
                success += 1
            else:
                _add_log(f"❌ [{idx+1}/{total}] {name} 发送失败")
                fail += 1

        except KeyError as e:
            _add_log(f"❌ [{idx+1}/{total}] {name} 文件未找到: {e}")
            fail += 1
        except Exception as e:
            _add_log(f"❌ [{idx+1}/{total}] {name} 发送失败: {e}")
            fail += 1

        # 间隔等待
        if idx < total - 1 and not state["stop_requested"]:
            time.sleep(2)

    # 完成
    if not state["stop_requested"]:
        state["result"] = "completed"
    _add_log(f"═══════════════════════")
    _add_log(f"发送完成: 成功 {success} / 失败 {fail} / 共 {total}")
    state["sending"] = False

    pythoncom.CoUninitialize()


# ══════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import webbrowser
    import threading

    print("=" * 50)
    print("  微信文件群发工具 — Web 版")
    print("  发送模式: 键盘模拟 (适用于所有微信版本)")
    print("  浏览器即将打开: http://localhost:5890")
    print("  关闭此窗口即可退出程序")
    print("=" * 50)

    # 启动后自动打开浏览器
    def _open_browser():
        time.sleep(1.5)
        webbrowser.open('http://localhost:5890')

    threading.Thread(target=_open_browser, daemon=True).start()

    app.run(host='127.0.0.1', port=5890, debug=False)
