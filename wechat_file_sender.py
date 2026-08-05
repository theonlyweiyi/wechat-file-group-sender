#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
微信文件群发工具
==================
根据 Excel 映射表（文件名关键词 → 微信昵称），自动将文件夹中的文件
匹配并发送给对应的微信联系人。

依赖:
  - wxauto4 (微信4.0客户端自动化)
  - openpyxl (读取Excel映射表)

使用前:
  1. 确保微信PC版已登录且窗口未最小化
  2. 准备Excel映射表（第一列:文件名关键词, 第二列:微信昵称）
  3. 准备存放待发送文件的文件夹

作者: WorkBuddy
日期: 2026-08-03
"""

import os
import sys
import time
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

# ─── 依赖检测 ───────────────────────────────────────────────
try:
    from openpyxl import load_workbook
except ImportError:
    print("缺少依赖 openpyxl，请运行: pip install openpyxl")
    sys.exit(1)

WXAUTO_VERSION = None
try:
    from wxauto4 import WeChat as WeChat4
    WXAUTO_VERSION = "4"
except ImportError:
    WeChat4 = None
    try:
        from wxauto import WeChat as WeChat3
        WXAUTO_VERSION = "3"
    except ImportError:
        WeChat3 = None


class WeChatFileSenderApp:
    """微信文件群发工具主界面"""

    # ── 颜色主题 ──
    COLOR_BG = "#f5f6f9"
    COLOR_CARD = "#ffffff"
    COLOR_PRIMARY = "#07c160"       # 微信绿
    COLOR_PRIMARY_DARK = "#06ad56"
    COLOR_DANGER = "#e74c3c"
    COLOR_TEXT = "#333333"
    COLOR_TEXT_LIGHT = "#888888"
    COLOR_BORDER = "#e0e0e0"
    COLOR_SUCCESS = "#07c160"
    COLOR_FAIL = "#e74c3c"
    COLOR_WARN = "#f39c12"

    def __init__(self, root):
        self.root = root
        self.root.title("微信文件群发工具")
        self.root.geometry("980x780")
        self.root.minsize(880, 680)
        self.root.configure(bg=self.COLOR_BG)

        # ── 状态变量 ──
        self.folder_path = tk.StringVar()
        self.excel_path = tk.StringVar()
        self.enable_message = tk.BooleanVar(value=False)
        self.message_text = tk.StringVar()
        self.delay_seconds = tk.DoubleVar(value=2.0)
        self.matched_data = []          # [(keyword, wechat_name, [file_paths])]
        self.is_sending = False
        self.should_stop = False
        self.wechat = None

        self._build_ui()
        self._check_wxauto()

    # ══════════════════════════════════════════════════════════
    #  UI 构建
    # ══════════════════════════════════════════════════════════

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")

        # ── 主容器 ──
        main_frame = tk.Frame(self.root, bg=self.COLOR_BG)
        main_frame.pack(fill="both", expand=True, padx=16, pady=12)

        # ── 标题栏 ──
        title_frame = tk.Frame(main_frame, bg=self.COLOR_BG)
        title_frame.pack(fill="x", pady=(0, 12))
        tk.Label(
            title_frame, text="📁  微信文件群发工具",
            font=("Microsoft YaHei UI", 18, "bold"),
            fg=self.COLOR_TEXT, bg=self.COLOR_BG
        ).pack(side="left")
        tk.Label(
            title_frame, text="按关键词匹配文件，自动发送给对应微信联系人",
            font=("Microsoft YaHei UI", 10),
            fg=self.COLOR_TEXT_LIGHT, bg=self.COLOR_BG
        ).pack(side="left", padx=(12, 0), pady=(6, 0))

        # ── 配置区 ──
        config_card = self._card_frame(main_frame)
        config_card.pack(fill="x", pady=(0, 10))

        # 文件夹路径
        row = 0
        tk.Label(config_card, text="文件夹路径", font=("Microsoft YaHei UI", 10, "bold"),
                 fg=self.COLOR_TEXT, bg=self.COLOR_CARD).grid(
            row=row, column=0, sticky="w", padx=16, pady=(14, 6))
        folder_entry = tk.Entry(config_card, textvariable=self.folder_path,
                                font=("Microsoft YaHei UI", 10), bd=1, relief="solid",
                                highlightthickness=0)
        folder_entry.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=(14, 6))
        tk.Button(config_card, text="浏览…", command=self._browse_folder,
                  font=("Microsoft YaHei UI", 9), width=8,
                  bg=self.COLOR_CARD, fg=self.COLOR_TEXT, bd=1, relief="solid",
                  cursor="hand2").grid(row=row, column=2, padx=(0, 16), pady=(14, 6))

        # 映射表文件
        row = 1
        tk.Label(config_card, text="映射表(Excel)", font=("Microsoft YaHei UI", 10, "bold"),
                 fg=self.COLOR_TEXT, bg=self.COLOR_CARD).grid(
            row=row, column=0, sticky="w", padx=16, pady=6)
        excel_entry = tk.Entry(config_card, textvariable=self.excel_path,
                               font=("Microsoft YaHei UI", 10), bd=1, relief="solid",
                               highlightthickness=0)
        excel_entry.grid(row=row, column=1, sticky="ew", padx=(0, 8), pady=6)
        tk.Button(config_card, text="浏览…", command=self._browse_excel,
                  font=("Microsoft YaHei UI", 9), width=8,
                  bg=self.COLOR_CARD, fg=self.COLOR_TEXT, bd=1, relief="solid",
                  cursor="hand2").grid(row=row, column=2, padx=(0, 16), pady=6)

        # 附带消息
        row = 2
        msg_frame = tk.Frame(config_card, bg=self.COLOR_CARD)
        msg_frame.grid(row=row, column=0, columnspan=3, sticky="ew", padx=16, pady=(6, 4))
        tk.Checkbutton(msg_frame, text="附带消息", variable=self.enable_message,
                       font=("Microsoft YaHei UI", 10, "bold"), fg=self.COLOR_TEXT,
                       bg=self.COLOR_CARD, activebackground=self.COLOR_CARD,
                       selectcolor=self.COLOR_CARD, bd=0,
                       command=self._toggle_message).pack(side="left")
        tk.Label(msg_frame, text="(支持 {name} 替换为联系人名)",
                 font=("Microsoft YaHei UI", 8), fg=self.COLOR_TEXT_LIGHT,
                 bg=self.COLOR_CARD).pack(side="left", padx=(8, 0))

        self.msg_entry = tk.Entry(config_card, textvariable=self.message_text,
                                  font=("Microsoft YaHei UI", 10), bd=1, relief="solid",
                                  highlightthickness=0, state="disabled")
        self.msg_entry.grid(row=3, column=0, columnspan=3, sticky="ew",
                            padx=16, pady=(0, 6))

        # 发送间隔
        row = 4
        delay_frame = tk.Frame(config_card, bg=self.COLOR_CARD)
        delay_frame.grid(row=row, column=0, columnspan=3, sticky="w", padx=16, pady=(0, 14))
        tk.Label(delay_frame, text="每次发送间隔", font=("Microsoft YaHei UI", 10, "bold"),
                 fg=self.COLOR_TEXT, bg=self.COLOR_CARD).pack(side="left")
        tk.Spinbox(delay_frame, from_=0.5, to=30, increment=0.5, width=5,
                   textvariable=self.delay_seconds, font=("Microsoft YaHei UI", 10),
                   bd=1, relief="solid").pack(side="left", padx=(8, 4))
        tk.Label(delay_frame, text="秒  (建议≥2秒，避免触发风控)",
                 font=("Microsoft YaHei UI", 9), fg=self.COLOR_TEXT_LIGHT,
                 bg=self.COLOR_CARD).pack(side="left", padx=(4, 0))

        config_card.grid_columnconfigure(1, weight=1)

        # ── 操作按钮区 ──
        btn_frame = tk.Frame(main_frame, bg=self.COLOR_BG)
        btn_frame.pack(fill="x", pady=(0, 10))

        self.preview_btn = tk.Button(
            btn_frame, text="🔍  预览匹配结果", command=self._preview_matches,
            font=("Microsoft YaHei UI", 11, "bold"), width=16,
            bg=self.COLOR_CARD, fg=self.COLOR_TEXT, bd=1, relief="solid",
            cursor="hand2", activebackground="#e8e8e8")
        self.preview_btn.pack(side="left", padx=(0, 10))

        self.send_btn = tk.Button(
            btn_frame, text="🚀  开始发送", command=self._start_sending,
            font=("Microsoft YaHei UI", 11, "bold"), width=14,
            bg=self.COLOR_PRIMARY, fg="white", bd=0, relief="flat",
            cursor="hand2", activebackground=self.COLOR_PRIMARY_DARK)
        self.send_btn.pack(side="left", padx=(0, 10))

        self.stop_btn = tk.Button(
            btn_frame, text="⏹  停止", command=self._stop_sending,
            font=("Microsoft YaHei UI", 11, "bold"), width=10,
            bg=self.COLOR_DANGER, fg="white", bd=0, relief="flat",
            cursor="hand2", activebackground="#c0392b", state="disabled")
        self.stop_btn.pack(side="left")

        self.stat_label = tk.Label(btn_frame, text="", font=("Microsoft YaHei UI", 9),
                                   fg=self.COLOR_TEXT_LIGHT, bg=self.COLOR_BG)
        self.stat_label.pack(side="right")

        # ── 匹配预览表 ──
        preview_card = self._card_frame(main_frame)
        preview_card.pack(fill="both", expand=True, pady=(0, 10))

        tk.Label(preview_card, text="📋  匹配预览", font=("Microsoft YaHei UI", 11, "bold"),
                 fg=self.COLOR_TEXT, bg=self.COLOR_CARD).pack(
            anchor="w", padx=16, pady=(10, 4))

        # Treeview
        tree_frame = tk.Frame(preview_card, bg=self.COLOR_CARD)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        columns = ("keyword", "name", "files", "status")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=8)
        self.tree.heading("keyword", text="文件名关键词")
        self.tree.heading("name", text="微信昵称")
        self.tree.heading("files", text="匹配到的文件")
        self.tree.heading("status", text="状态")
        self.tree.column("keyword", width=140, anchor="w")
        self.tree.column("name", width=120, anchor="w")
        self.tree.column("files", width=400, anchor="w")
        self.tree.column("status", width=80, anchor="center")

        tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side="left", fill="both", expand=True)
        tree_scroll.pack(side="right", fill="y")

        # Treeview 样式
        style.configure("Treeview", font=("Microsoft YaHei UI", 9), rowheight=28)
        style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 9, "bold"))
        self.tree.tag_configure("success", foreground=self.COLOR_SUCCESS)
        self.tree.tag_configure("fail", foreground=self.COLOR_FAIL)
        self.tree.tag_configure("warn", foreground=self.COLOR_WARN)
        self.tree.tag_configure("pending", foreground=self.COLOR_TEXT_LIGHT)

        # ── 日志区 ──
        log_card = self._card_frame(main_frame)
        log_card.pack(fill="both", expand=True, pady=(0, 10))

        tk.Label(log_card, text="📝  发送日志", font=("Microsoft YaHei UI", 11, "bold"),
                 fg=self.COLOR_TEXT, bg=self.COLOR_CARD).pack(
            anchor="w", padx=16, pady=(10, 4))

        log_frame = tk.Frame(log_card, bg=self.COLOR_CARD)
        log_frame.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        self.log_text = tk.Text(log_frame, font=("Consolas", 9), height=10,
                                bd=1, relief="solid", highlightthickness=0,
                                bg="#fafafa", fg=self.COLOR_TEXT, state="disabled",
                                wrap="word")
        log_scroll = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side="left", fill="both", expand=True)
        log_scroll.pack(side="right", fill="y")

        # ── 进度条 ──
        self.progress = ttk.Progressbar(main_frame, mode="determinate")
        self.progress.pack(fill="x", pady=(0, 4))

        # ── 底部状态栏 ──
        status_bar = tk.Frame(main_frame, bg=self.COLOR_BG)
        status_bar.pack(fill="x", pady=(0, 2))
        self.status_var = tk.StringVar(value="就绪")
        tk.Label(status_bar, textvariable=self.status_var,
                 font=("Microsoft YaHei UI", 9), fg=self.COLOR_TEXT_LIGHT,
                 bg=self.COLOR_BG).pack(side="left")
        wx_info = f"wxauto{'4' if WXAUTO_VERSION == '4' else ''} v{WXAUTO_VERSION or '未安装'}"
        tk.Label(status_bar, text=wx_info,
                 font=("Microsoft YaHei UI", 9), fg=self.COLOR_TEXT_LIGHT,
                 bg=self.COLOR_BG).pack(side="right")

    def _card_frame(self, parent):
        """创建白色卡片容器"""
        frame = tk.Frame(parent, bg=self.COLOR_CARD, bd=0, relief="flat",
                         highlightbackground=self.COLOR_BORDER, highlightthickness=1)
        return frame

    # ══════════════════════════════════════════════════════════
    #  交互逻辑
    # ══════════════════════════════════════════════════════════

    def _check_wxauto(self):
        """检查 wxauto 是否可用"""
        if WXAUTO_VERSION is None:
            self._log("⚠ wxauto 未安装，无法发送。请运行: pip install wxauto4")
            self._set_status("wxauto 未安装")
            self.send_btn.configure(state="disabled")

    def _toggle_message(self):
        """切换消息输入框的可用状态"""
        if self.enable_message.get():
            self.msg_entry.configure(state="normal")
        else:
            self.msg_entry.configure(state="disabled")

    def _browse_folder(self):
        path = filedialog.askdirectory(title="选择包含待发送文件的文件夹")
        if path:
            self.folder_path.set(path)
            self._log(f"已选择文件夹: {path}")

    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="选择Excel映射表",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if path:
            self.excel_path.set(path)
            self._log(f"已选择映射表: {path}")

    # ── 匹配逻辑 ───────────────────────────────────────────

    def _read_excel_mapping(self, filepath):
        """读取Excel映射表，返回 [(keyword, wechat_name), ...]"""
        mappings = []
        ext = os.path.splitext(filepath)[1].lower()

        if ext in (".xlsx", ".xls"):
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(min_row=1, values_only=True)):
                if i == 0:
                    # 跳过表头
                    continue
                if not row or len(row) < 2:
                    continue
                keyword = str(row[0]).strip() if row[0] else ""
                name = str(row[1]).strip() if row[1] else ""
                if keyword and name:
                    mappings.append((keyword, name))
            wb.close()
        elif ext == ".csv":
            import csv
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i == 0:
                        continue
                    if len(row) < 2:
                        continue
                    keyword = row[0].strip()
                    name = row[1].strip()
                    if keyword and name:
                        mappings.append((keyword, name))
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

        return mappings

    def _scan_folder(self, folder):
        """扫描文件夹，返回所有文件的 [(filename, abspath), ...]"""
        files = []
        for name in os.listdir(folder):
            full_path = os.path.join(folder, name)
            if os.path.isfile(full_path):
                files.append((name, full_path))
        return files

    def _match_files(self, mappings, files):
        """关键词匹配：文件名包含关键词即匹配（不区分大小写）"""
        matched = []
        for keyword, name in mappings:
            matched_files = []
            kw_lower = keyword.lower()
            for filename, filepath in files:
                if kw_lower in filename.lower():
                    matched_files.append(filepath)
            matched.append((keyword, name, matched_files))
        return matched

    def _preview_matches(self):
        """预览匹配结果"""
        # 验证输入
        folder = self.folder_path.get().strip()
        excel = self.excel_path.get().strip()

        if not folder:
            messagebox.showwarning("提示", "请先选择文件夹路径")
            return
        if not excel:
            messagebox.showwarning("提示", "请先选择映射表文件")
            return
        if not os.path.isdir(folder):
            messagebox.showerror("错误", f"文件夹不存在: {folder}")
            return
        if not os.path.isfile(excel):
            messagebox.showerror("错误", f"映射表文件不存在: {excel}")
            return

        # 读取映射表
        try:
            mappings = self._read_excel_mapping(excel)
        except Exception as e:
            messagebox.showerror("错误", f"读取映射表失败:\n{e}")
            return

        if not mappings:
            messagebox.showwarning("提示", "映射表中没有有效数据\n(第一列为关键词，第二列为微信昵称)")
            return

        # 扫描文件夹
        files = self._scan_folder(folder)

        # 匹配
        self.matched_data = self._match_files(mappings, files)

        # 清空表格
        for item in self.tree.get_children():
            self.tree.delete(item)

        # 填充表格
        total_files = 0
        matched_count = 0
        no_match_count = 0
        for keyword, name, file_list in self.matched_data:
            if file_list:
                file_display = "\n".join(os.path.basename(f) for f in file_list)
                self.tree.insert("", "end", values=(keyword, name, file_display, "待发送"),
                                 tags=("pending",))
                total_files += len(file_list)
                matched_count += 1
            else:
                self.tree.insert("", "end", values=(keyword, name, "⚠ 未匹配到文件", "跳过"),
                                 tags=("warn",))
                no_match_count += 1

        self.stat_label.configure(
            text=f"共 {len(self.matched_data)} 条映射 | "
                 f"匹配成功 {matched_count} 条 | 未匹配 {no_match_count} 条 | "
                 f"待发文件 {total_files} 个"
        )
        self._log(f"预览完成: {len(self.matched_data)} 条映射, 匹配到 {total_files} 个文件")
        self._set_status("预览完成，点击「开始发送」执行")

    # ── 发送逻辑 ───────────────────────────────────────────

    def _start_sending(self):
        """开始发送"""
        if not self.matched_data:
            # 自动预览
            self._preview_matches()
            if not self.matched_data:
                messagebox.showwarning("提示", "没有可发送的数据，请先预览匹配结果")
                return

        # 过滤出有文件的条目
        send_list = [(kw, name, files) for kw, name, files in self.matched_data if files]
        if not send_list:
            messagebox.showwarning("提示", "没有匹配到任何文件，无法发送")
            return

        # 确认
        total = len(send_list)
        if not messagebox.askyesno("确认发送",
                f"即将向 {total} 个联系人发送文件。\n\n"
                f"⚠ 请确保:\n"
                f"  1. 微信PC版已登录且窗口未最小化\n"
                f"  2. 映射表中的微信昵称与微信一致\n"
                f"  3. 发送过程中不要操作微信窗口\n\n"
                f"确认开始发送？"):
            return

        # 切换按钮状态
        self.is_sending = True
        self.should_stop = False
        self.send_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.preview_btn.configure(state="disabled")

        # 在后台线程中执行发送
        thread = threading.Thread(target=self._send_worker, args=(send_list,), daemon=True)
        thread.start()

    def _stop_sending(self):
        """停止发送"""
        self.should_stop = True
        self._log("⏹ 用户请求停止发送，将在当前联系人发送完成后停止...")
        self._set_status("正在停止...")

    def _send_worker(self, send_list):
        """发送工作线程"""
        total = len(send_list)
        self.progress_configure(maximum=total, value=0)

        # 初始化微信
        self._log_threadsafe("正在连接微信客户端...")
        self._set_status_threadsafe("连接微信中...")

        try:
            if WXAUTO_VERSION == "4":
                self.wechat = WeChat4()
            elif WXAUTO_VERSION == "3":
                self.wechat = WeChat3()
            else:
                self._log_threadsafe("❌ wxauto 未安装")
                self._finish_sending()
                return
            self._log_threadsafe("✅ 微信连接成功")
        except Exception as e:
            self._log_threadsafe(f"❌ 微信连接失败: {e}")
            self._log_threadsafe("请确保微信PC版已登录且窗口未最小化")
            self._finish_sending()
            return

        delay = self.delay_seconds.get()
        send_msg = self.enable_message.get()
        msg_template = self.message_text.get().strip()

        success_count = 0
        fail_count = 0

        for idx, (keyword, name, file_list) in enumerate(send_list):
            if self.should_stop:
                self._log_threadsafe(f"⏹ 已停止发送 (完成 {idx}/{total})")
                break

            self._set_status_threadsafe(f"发送中: {name} ({idx+1}/{total})")
            self._update_tree_status(idx, "发送中…", "pending")

            try:
                # 发送消息（如果启用）
                if send_msg and msg_template:
                    msg = msg_template.replace("{name}", name)
                    self.wechat.SendMsg(msg, name)
                    self._log_threadsafe(f"  → 消息已发送给 {name}")

                time.sleep(0.5)

                # 发送文件
                if len(file_list) == 1:
                    self.wechat.SendFiles(file_list[0], name)
                else:
                    self.wechat.SendFiles(file_list, name)

                file_names = ", ".join(os.path.basename(f) for f in file_list)
                self._log_threadsafe(
                    f"✅ [{idx+1}/{total}] {name} ← {file_names}")
                self._update_tree_status(idx, "✅ 成功", "success")
                success_count += 1

            except Exception as e:
                self._log_threadsafe(
                    f"❌ [{idx+1}/{total}] {name} 发送失败: {e}")
                self._update_tree_status(idx, "❌ 失败", "fail")
                fail_count += 1

            # 更新进度条
            self.progress_configure_threadsafe(value=idx + 1)

            # 间隔等待（最后一个不需要）
            if idx < total - 1 and not self.should_stop:
                self._set_status_threadsafe(f"等待 {delay}s…")
                time.sleep(delay)

        # 完成
        self._log_threadsafe(
            f"\n════════════════════════════════\n"
            f"发送完成: 成功 {success_count} / 失败 {fail_count} / 共 {total}\n"
            f"════════════════════════════════")
        self._finish_sending()

    def _finish_sending(self):
        """发送完成后的清理"""
        def _reset():
            self.is_sending = False
            self.should_stop = False
            self.send_btn.configure(state="normal")
            self.stop_btn.configure(state="disabled")
            self.preview_btn.configure(state="normal")
            self._set_status("就绪")
        self.root.after(0, _reset)

    # ── 线程安全的UI更新方法 ─────────────────────────────────

    def _log(self, message):
        """向日志区写入消息（主线程调用）"""
        self._append_log(message)

    def _log_threadsafe(self, message):
        """向日志区写入消息（子线程安全调用）"""
        self.root.after(0, lambda: self._append_log(message))

    def _append_log(self, message):
        """实际写入日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _set_status(self, text):
        self.status_var.set(text)

    def _set_status_threadsafe(self, text):
        self.root.after(0, lambda: self.status_var.set(text))

    def _update_tree_status(self, index, status, tag):
        """更新表格中某行的状态"""
        def _update():
            children = self.tree.get_children()
            # 只更新有文件且状态为"待发送"的行
            send_idx = 0
            for child in children:
                vals = self.tree.item(child, "values")
                if vals[3] in ("待发送", "发送中…", "✅ 成功", "❌ 失败"):
                    if send_idx == index:
                        self.tree.item(child, values=(vals[0], vals[1], vals[2], status),
                                       tags=(tag,))
                        break
                    send_idx += 1
        self.root.after(0, _update)

    def progress_configure(self, **kwargs):
        self.progress.configure(**kwargs)

    def progress_configure_threadsafe(self, **kwargs):
        self.root.after(0, lambda: self.progress.configure(**kwargs))


# ═════════════════════════════════════════════════════════════
#  启动
# ═════════════════════════════════════════════════════════════

def main():
    root = tk.Tk()
    app = WeChatFileSenderApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
